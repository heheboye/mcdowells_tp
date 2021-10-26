from openpyxl import Workbook, load_workbook
import colorama
import time
import os

colorama.init()

white = "\x1b[1;37;40m" 
red = "\x1b[1;31;222m"

custom_date = time.strftime("%d %m %Y")
raw_date = time.asctime()
headers = [["Cliente", "Fecha", "Combo S", "Combo D", "Combo T", "Flurby", "Total"]]
rows = []
total_sales = []

### FUNCTIONS ###

def verify(user_input):
    while user_input == "":
        print(f"""\n{white}El campo no puede estar vacio.
        Intente nuevamente: """, end = "")
        user_input = input(f"{red}")
    return user_input

def int_convert(user_input):
    while user_input.isdecimal() == False or user_input == "":
        print(f"""\n{white}El campo no puede estar vacio y solo se aceptan numeros enteros.
        Intente nuevamente: """, end = "")
        user_input = input(f"{red}")
    user_input = int(user_input)
    return user_input

def emp_in(emp):
    f = open(f"Registro {custom_date}.txt", "a")
    f.write(f"IN {raw_date} Encargad@ {emp}\n")
    f.close()
    
def emp_out(emp):
    f = open(f"Registro {custom_date}.txt", "a")
    f.write(f"OUT {raw_date} Encargad@ {emp} ${sum(total_sales)}\n")
    f.write("#"*50)
    f.close()

def reg_sale():
    if os.path.exists(f"Registro {time.strftime('%d %m %Y')}.xlsx"):
        wb = load_workbook(filename = f"Registro {time.strftime('%d %m %Y')}.xlsx")
        ws = wb.active
        for row in rows:
            ws.append(row)
        wb.save(f"Registro {time.strftime('%d %m %Y')}.xlsx")
    else:
        wb = Workbook()
        ws = wb.active
        for head in headers:
            ws.append(head)
        for row in rows:
            ws.append(row)
        wb.save(f"Registro {time.strftime('%d %m %Y')}.xlsx")

def clear_console():
    if os.name == "nt":
        os.system("cls")
    else:
        os.system("clear")

### PRIMER MENU ###

print(f"\n{white}Bienvenido a McDowell's\n", end = "")
print("\nIngrese su nombre encargad@: ", end = "")
encargado = input(f"{red}")
encargado = verify(encargado)
emp_in(encargado)

###################

while True:
    print(f"""{white}
McDowell's
Recuerda que siempre hay que recibir al cliente con una sonrisa :)

1 – Ingreso de nuevo pedido
2 – Cambio de turno
3 – Apagar sistema

""", end = "")
    option = input(">>> ")
    option = int_convert(option)
    clear_console()
    if option == 1:
        clear_console()
        print(f"{white}Nombre del cliente: ", end = "")
        customer = input(f"{red}")
        customer = verify(customer)
        print(f"{white}Ingrese cantidad Combo S: ", end = "")
        can_cs = input(f"{red}")
        can_cs = int_convert(can_cs)
        print(f"{white}Ingrese cantidad Combo D: ", end = "")
        can_cd = input(f"{red}")
        can_cd = int_convert(can_cd)
        print(f"{white}Ingrese cantidad Combo T: ", end = "")
        can_ct = input(f"{red}")
        can_ct = int_convert(can_ct)
        print(f"{white}Ingrese cantidad Flurby: ", end = "")
        can_p = input(f"{red}")
        can_p = int_convert(can_p)
        total = (can_cs * 650) + (can_cd * 700) + (can_ct * 800) + (can_p * 250)
        print(f"\n{white}Total $ {total}")
        print(f"Abona con $ ", end = "")
        pago = input(f"{red}")
        pago = int_convert(pago)
        clear_console()
        print(f"{white}El vuelto de {red}{customer}{white} es de $ {red}{pago - total}")
        total_sales.append(total)
        rows.append([customer, raw_date, can_cs, can_cd, can_ct, can_p, total])
        reg_sale()
        rows = []
    if option == 2:
        emp_out(encargado)
        total_sales = []
        print(f"\n{white}Bienvenido a McDowell's\n", end = "")
        print("\nIngrese su nombre encargad@: ", end = "")
        encargado = input(f"{red}")
        encargado = verify(encargado)
        emp_in(encargado)
        clear_console()
    if option == 3:
        emp_out(encargado)
        print("\nGuargado informacion del encargado anterior...")
        time.sleep(3)
        clear_console()
        print("\nPrograma cerrado exitosamente.", end = "")
        break